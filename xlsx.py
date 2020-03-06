from openpyxl import load_workbook
import re
workbook = load_workbook('harnexdata.xlsx')
worksheet = workbook.get_sheet_by_name('Sheet1')
sheet_cells = []
grp1=[(1,5,6,'v','H-78'),(2,5,6,'v','H-79'),(3,10,4,'C','C-76'),(4,4,7,'D','C-79')]
grp2=[(1,2,3,4,5,6),(7,8,9,10,11,12,13),(14,15,16,17,18,19,20,21,22,23),(24,25,16,17,18,19,20,21,22,23)]

for rows in worksheet.iter_rows():
    row_cells = []
    for cell in rows:
        row_cells.append(str(cell.value))
	#print(cell.value)
    sheet_cells.append(tuple(row_cells))
##    print(sheet_cells)
    
def harnex(sheet_cells,grp1,grp2):
##    counter=1
##    temp_var = []
##    temp1=[]
##    temp2=[]
    
##    ress=[]
##    net=[]
    for i in range (20):
##        if(re.search(regex,str(i))):
##            print(i)
            for sheet_rowtup in sheet_cells:                                                ##sheet_rowtup is exel tuple
                if (sheet_rowtup[0]==str(i)):
##                    print(sheet_rowtup)
                    conn_name1=sheet_rowtup[2]                                              ##conn_name1 and conn_name2 are connector names
                    conn_name2=sheet_rowtup[4]
                    print(conn_name1)
                    print(conn_name2)
                    print(sheet_rowtup)
##                    for j in range (len(grp1)):
                        
##                            skip = True
                    grp1_tup1 = []
                    grp2_tup1 = []
                    grp1_tup2 = []
                    grp2_tup2 = []
                    loop_cnt=0
                    for j in range (len(grp1)):
                        temp = grp1[j]
                        
                        if(conn_name1==conn_name2):
                            if (conn_name1 in grp1[j]):
                                grp1_tup1 = grp1[j]
                                grp2_tup1 = grp2[j]
##                                print(grp2_tup1)
##                                print(grp1[j])
##                                print(grp2[j])
                                if(loop_cnt==(len(grp1)-1)):
                                    print('grp_1',grp1_tup1,conn_name1,'grp2',grp2_tup1,conn_name2)
                        else:
                            if (conn_name1 in grp1[j]):
                                grp1_tup1 = grp1[j]
                                grp2_tup1 = grp2[j]
##                                print(grp2[j])
##                                print(grp1[j])
    ##                            skip = False
                            elif (conn_name2 in grp1[j]):
                                grp1_tup2 = grp1[j]
                                grp2_tup2 = grp2[j]
                            if(loop_cnt==(len(grp1)-1)):
                                print('grp_1',grp1_tup1,grp1_tup2,'conn1 -',conn_name1,'conn2 -',conn_name2,'grp2',grp2_tup1,grp2_tup2)
                        loop_cnt=loop_cnt+1
                    net=[]
                    if((len(grp2_tup1)>0)and(len(grp2_tup2)>0)):
                        vall = int(sheet_rowtup[3])
                        if(vall==0):
                            temp21 = 0
                        else:
                            temp21 = grp2_tup1[(vall-1)]
                        vall = int(sheet_rowtup[5])
                        if(vall==0):
                            temp22 = 0
                        else:
                            temp22 = grp2_tup2[(vall-1)]
                        net.append([temp21,temp22])
##                        net.append([grp2_tup1[int(sheet_rowtup[3])-1],grp2_tup2[int(sheet_rowtup[5])-1]])
                    elif((len(grp2_tup1)>0)):
                        vall = int(sheet_rowtup[3])
                        if(vall==0):
                            temp21 = 0
                        else:
                            temp21 = grp2_tup1[(vall-1)]
                        vall = int(sheet_rowtup[5])
                        if(vall==0):
                            temp22 = 0
                        else:
                            temp22 = grp2_tup1[(vall-1)]
                        net.append([temp21,temp22])
##                        net.append([grp2_tup1[int(sheet_rowtup[3])-1],grp2_tup1[int(sheet_rowtup[5])-1]])
                    elif((len(grp2_tup2)>0)):
                        vall = int(sheet_rowtup[3])
                        if(vall==0):
                            temp21 = 0
                        else:
                            temp21 = grp2_tup2[(vall-1)]
                        vall = int(sheet_rowtup[5])
                        if(vall==0):
                            temp22 = 0
                        else:
                            temp22 = grp2_tup2[(vall-1)]
                        net.append([temp21,temp22])
##                        net.append([grp2_tup2[int(sheet_rowtup[3])-1],grp2_tup2[int(sheet_rowtup[5])-1]])
##                    elif(conn_name1==conn_name2):
##                        net.append([grp2_tup2[int(sheet_rowtup[3])-1],grp2_tup2[int(sheet_rowtup[5])-1]],[grp2_tup1[int(sheet_rowtup[3])-1]])
                    print(net)
##                            skip = Falsez
##                        elif(not((conn_name1 in temp)and(conn_name2 in temp))):
##                            print('skip',grp1_tup1)
##                            grp1_tup1=[]
##                            grp1_tup2=[]
##                        grp1_tup1 = grp1[j]                                                 ##grp1_tup1 is grp1 tuple 
##                        res=[]
##                        if (len(grp1_tup1)>1):
##                            if(conn_name1==conn_name2):
    ##                            if (conn_name1 == grp1_tup1[4]):
##                                grp2_tup1 = grp2[grp1_tup1[0]-1]
    ##                                print('grp_1',grp1_tup1,conn_name1,grp2_tup1)
    ##                                if(conn_name1==conn_name2):
    ##                                res.append(grp2_tup1)
##                                print('grp_1',grp1_tup1,conn_name1,'grp2',grp2_tup1)
##                            else:
    ##                            if (conn_name2 == grp1_tup1[4]):
    ##                                if(not(conn_name1==conn_name2)):
##                                grp2_tup1 = grp2[grp1_tup1[0]-1]
##                                grp2_tup2 = grp2[grp1_tup2[0]-1]
    ##                                res.append(grp2_tup1)
##                                print('grp_1',grp1_tup1,conn_name1,'grp2',grp2_tup1,conn_name2,grp2_tup2)
##                    print(res)



##                            print(grp2_tup1[int(sheet_rowtup[3])-1],grp2_tup1[int(sheet_rowtup[5])-1])
                            
##                            net.append(grp2_tup1[int(sheet_rowtup[3])-1])
##                            net.append(grp2_tup1[int(sheet_rowtup[5])-1])
##                    print(net)
####                            print(grp1_tup1[0])
##                            for x in grp2:
##                                if (grp1_tup1[0]==x):
####                                    temp_var.append(str(x))
##                                    temp1=x
##                                    res.append(x[int(sheet_rowtup[3])-1])
####                                    print(x)
##                        if conn_name2 in grp1_tup1:
####                            print(t)                   #tuple using part name
##                            for y in grp2:
##                                #print(len(z),type(t[1]),t[1])
##                                if (grp1_tup1[0]==y):
##        ##                                temp_var.append(str(y))
##                                    temp2=y
##                                    res.append(y[int(sheet_rowtup[5])-1])
####                                    print(y)     #no of wiring pins in grp2
####                        print('res',res)
##                        ress.append(tuple(res))
####    counter=counter+1
##
##    return ([temp1,temp2],ress)
##hrn,res = harnex(sheet_cells,grp1,grp2)
harnex(sheet_cells,grp1,grp2)
##print('hrn',hrn)
##print('res',res)
