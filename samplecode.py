from openpyxl import load_workbook
workbook = load_workbook('harnexdata.xlsx')
worksheet = workbook.get_sheet_by_name('Sheet1')
sheet_cells = []
grp1=[(1,6,2,'A','H-78'),(2,7,3,'A2','H-79')]
grp2=[(1,2,3,4,5,6),(7,8,9,10,11,12,13)]

for rows in worksheet.iter_rows():
    row_cells = []
    for cell in rows:
        row_cells.append(str(cell.value))
	#print(cell.value)
    sheet_cells.append(tuple(row_cells))
##print(sheet_cells)
temp_code = []
res = []
ress = []
def harnex(sheet_cells,grp1,grp2):
    for i in range (20):
        for sheet_rowtup in sheet_cells:
            if (sheet_rowtup[0]==str(i)):
                conn_name1=sheet_rowtup[2]                                              
                conn_name2=sheet_rowtup[4]
##                print(conn_name1)
##                print(conn_name2)
                for j in range (len(grp1)):
                    grp1_tup1 = grp1[j]
                    print(grp1_tup1)
                    if conn_name1 in grp1_tup1:
                        print(conn_name1)                
                        for x in grp2:
                            if (grp1_tup1[1]==int(len(x))):
##                                temp_var.append(str(x))
                                temp1=x
                                res.append(x[int(sheet_rowtup[3])-1])
                                print(x)
                    if conn_name2 in grp1_tup1:
                        print(conn_name2)                   #tuple using part name
                        for y in grp2:
                            #print(len(z),type(t[1]),t[1])
                            if (grp1_tup1[1]==int(len(y))):
##                                temp_var.append(str(y))
                                temp2=y
                                res.append(y[int(sheet_rowtup[5])-1])
                                print(y)     #no of wiring pins in grp2
##                    print('res',res)
                    ress.append(tuple(res))
                




    return ([temp1,temp2],ress)
hrn,res = harnex(sheet_cells,grp1,grp2)
##harnex(sheet_cells,grp1,grp2)
##print('hrn',hrn)
##for i in range (10):
##	for j in sheet_cells:
##		if (j[1]==str(i)):
##			print(j)                                  #tuple using ckt no
##			for t in grp1:                             #search in grp1
##				if j[2] in t:
##					print(t)
##					for z in grp2:
##						if (t[0]==len(z)):
##							temp_code.append(tuple(z))
##							print(z)
##				if j[4] in t:
##					print(t)                   #tuple using part name
##					for y in grp2:
##						#print(len(z),type(t[1]),t[1])
##						if (int(t[0])==int(len(y))):
##							temp_code.append(tuple(y))
##							print(y)     #no of wiring pins in grp2
##
